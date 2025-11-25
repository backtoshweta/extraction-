"""
utils_v3_complete.py - Complete Enhanced Document Processing with Semantic Chunking
Features: Better chunking strategies, improved vector DB, BM25 integration, robust error handling
"""

import io
import os
import mimetypes
import tempfile
import concurrent.futures as cf
from pathlib import Path
from typing import Dict, List, Tuple
import re

import fitz  # PyMuPDF
import pytesseract
from PIL import Image
import docx2txt
from openpyxl import load_workbook

# LangChain imports for enhanced vector store creation
from langchain_community.vectorstores import Chroma
from langchain_ollama import OllamaEmbeddings
from langchain.retrievers import BM25Retriever, EnsembleRetriever
from langchain.schema import Document
import chromadb

# ================================================================== #
# 1. Enhanced Text Extraction                                       #
# ================================================================== #

_OCR_LANG = "eng"

def _pdf_extract_fast(file_bytes: bytes) -> str:
    """Enhanced PDF extraction with better OCR handling."""
    try:
        doc = fitz.open(stream=file_bytes, filetype="pdf")
        texts = []
        
        for page_num in range(doc.page_count):
            page = doc[page_num]
            text = page.get_text()
            
            if text.strip():
                texts.append(text)
            else:
                # Enhanced OCR for scanned pages
                try:
                    pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))  # Higher resolution
                    img_data = pix.tobytes("png")
                    img = Image.open(io.BytesIO(img_data))
                    
                    # Enhanced OCR configuration
                    custom_config = r'--oem 3 --psm 6 -c tessedit_char_whitelist=0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz.,!?@#$%^&*()_+-=[]{}|;:,.<>?/~` '
                    ocr_text = pytesseract.image_to_string(img, config=custom_config, lang=_OCR_LANG)
                    texts.append(ocr_text)
                except Exception as e:
                    print(f"OCR failed for page {page_num}: {e}")
                    texts.append("")
        
        doc.close()
        return "\n".join(texts)
    except Exception as e:
        print(f"PDF extraction error: {e}")
        return ""

def _image_extract(file_bytes: bytes) -> str:
    """Enhanced image OCR with preprocessing."""
    try:
        img = Image.open(io.BytesIO(file_bytes))
        
        # Preprocess image for better OCR
        img = img.convert('RGB')
        
        # Enhanced OCR config
        custom_config = r'--oem 3 --psm 6 -c tessedit_char_whitelist=0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz.,!?@#$%^&*()_+-=[]{}|;:,.<>?/~` '
        return pytesseract.image_to_string(img, config=custom_config, lang=_OCR_LANG)
    except Exception as e:
        print(f"Image OCR error: {e}")
        return ""

def _docx_extract(file_bytes: bytes, filename: str) -> str:
    """Enhanced DOCX extraction with structure preservation."""
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".docx") as tmp:
            tmp.write(file_bytes)
            tmp_path = tmp.name
        
        text = docx2txt.process(tmp_path)
        os.unlink(tmp_path)
        
        # Clean up excessive whitespace but preserve paragraph breaks
        text = re.sub(r'\n\s*\n', '\n\n', text)  # Normalize paragraph breaks
        text = re.sub(r'[ \t]+', ' ', text)      # Normalize spaces
        
        return text
    except Exception as e:
        print(f"DOCX extraction error: {e}")
        return ""

def _xlsx_extract(file_bytes: bytes, filename: str) -> str:
    """Enhanced XLSX extraction with better formatting."""
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(file_bytes)
            tmp_path = tmp.name
        
        wb = load_workbook(tmp_path, data_only=True)
        all_text = []
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_text = [f"=== SHEET: {sheet_name} ==="]
            
            # Extract with row/column structure
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                if any(cell is not None for cell in row):  # Skip empty rows
                    row_text = [str(cell) if cell is not None else "" for cell in row]
                    sheet_text.append(f"Row {row_idx}: " + " | ".join(row_text))
            
            all_text.append("\n".join(sheet_text))
        
        os.unlink(tmp_path)
        return "\n\n".join(all_text)
    except Exception as e:
        print(f"XLSX extraction error: {e}")
        return ""

def _txt_extract(file_bytes: bytes) -> str:
    """Enhanced text extraction with encoding detection."""
    encodings = ['utf-8', 'latin-1', 'ascii', 'utf-16']
    
    for encoding in encodings:
        try:
            return file_bytes.decode(encoding, errors='ignore')
        except (UnicodeDecodeError, LookupError):
            continue
    
    # Fallback
    return file_bytes.decode('utf-8', errors='replace')

# File type mapping
_EXTRACTORS = {
    ".pdf": _pdf_extract_fast,
    ".docx": _docx_extract,
    ".xlsx": _xlsx_extract,
    ".txt": _txt_extract,
    ".csv": _txt_extract,
    ".png": _image_extract,
    ".jpg": _image_extract,
    ".jpeg": _image_extract,
}

def extract_text_from_file(file_bytes: bytes, filename: str) -> str:
    """Extract text from any supported file type with enhanced processing."""
    ext = Path(filename).suffix.lower()
    
    if ext in _EXTRACTORS:
        extractor = _EXTRACTORS[ext]
        if ext in [".docx", ".xlsx"]:
            return extractor(file_bytes, filename)
        else:
            return extractor(file_bytes)
    
    # Fallback to MIME type detection
    mtype, _ = mimetypes.guess_type(filename)
    if mtype and mtype.startswith("image/"):
        return _image_extract(file_bytes)
    
    raise ValueError(f"Unsupported file type: {ext}")

# ================================================================== #
# 2. Enhanced Chunking Strategies                                   #
# ================================================================== #

def _semantic_chunk_text(text: str, max_tokens: int = 512, overlap_tokens: int = 50) -> List[str]:
    """
    Semantic chunking that preserves sentence and paragraph boundaries.
    Better than fixed-size chunking for maintaining context.
    """
    if not text.strip():
        return []
    
    # First, split by paragraphs
    paragraphs = re.split(r'\n\s*\n', text.strip())
    paragraphs = [p.strip() for p in paragraphs if p.strip()]
    
    if not paragraphs:
        return [text]
    
    chunks = []
    current_chunk = ""
    current_tokens = 0
    
    for paragraph in paragraphs:
        # Estimate tokens (rough: ~4 chars per token)
        para_tokens = len(paragraph) // 4
        
        # If paragraph alone exceeds max_tokens, split by sentences
        if para_tokens > max_tokens:
            sentences = _split_into_sentences(paragraph)
            for sentence in sentences:
                sent_tokens = len(sentence) // 4
                
                if current_tokens + sent_tokens > max_tokens and current_chunk:
                    # Finalize current chunk
                    chunks.append(current_chunk.strip())
                    
                    # Start new chunk with overlap
                    if overlap_tokens > 0:
                        overlap_text = _get_last_n_tokens(current_chunk, overlap_tokens)
                        current_chunk = overlap_text + " " + sentence
                        current_tokens = overlap_tokens + sent_tokens
                    else:
                        current_chunk = sentence
                        current_tokens = sent_tokens
                else:
                    current_chunk += " " + sentence if current_chunk else sentence
                    current_tokens += sent_tokens
        else:
            # Normal paragraph processing
            if current_tokens + para_tokens > max_tokens and current_chunk:
                # Finalize current chunk
                chunks.append(current_chunk.strip())
                
                # Start new chunk with overlap
                if overlap_tokens > 0:
                    overlap_text = _get_last_n_tokens(current_chunk, overlap_tokens)
                    current_chunk = overlap_text + "\n\n" + paragraph
                    current_tokens = overlap_tokens + para_tokens
                else:
                    current_chunk = paragraph
                    current_tokens = para_tokens
            else:
                current_chunk += "\n\n" + paragraph if current_chunk else paragraph
                current_tokens += para_tokens
    
    # Add final chunk
    if current_chunk.strip():
        chunks.append(current_chunk.strip())
    
    return chunks if chunks else [text]

def _split_into_sentences(text: str) -> List[str]:
    """Split text into sentences using regex."""
    # Basic sentence splitting (can be enhanced with NLTK if needed)
    sentences = re.split(r'(?<=[.!?])\s+', text)
    return [s.strip() for s in sentences if s.strip()]

def _get_last_n_tokens(text: str, n_tokens: int) -> str:
    """Get approximately the last n tokens from text."""
    words = text.split()
    # Rough estimate: 1.3 words per token
    n_words = int(n_tokens * 1.3)
    return " ".join(words[-n_words:]) if len(words) > n_words else text

def _hybrid_chunk_text(text: str, max_tokens: int = 512, overlap_tokens: int = 50) -> List[str]:
    """
    Hybrid chunking that combines semantic and fixed-size approaches.
    Falls back to semantic chunking if document has clear structure.
    """
    if not text.strip():
        return []
    
    # Check if text has clear paragraph structure
    paragraph_count = len(re.findall(r'\n\s*\n', text))
    sentence_count = len(re.findall(r'[.!?]+', text))
    
    # Use semantic chunking if text has good structure
    if paragraph_count > 2 or sentence_count > 10:
        return _semantic_chunk_text(text, max_tokens, overlap_tokens)
    else:
        # Fall back to improved fixed-size chunking
        return _improved_fixed_chunk(text, max_tokens, overlap_tokens)

def _improved_fixed_chunk(text: str, max_tokens: int = 512, overlap_tokens: int = 50) -> List[str]:
    """Improved fixed-size chunking that tries to break at sentence boundaries."""
    if not text.strip():
        return []
    
    # Estimate characters per token (roughly 4)
    max_chars = max_tokens * 4
    overlap_chars = overlap_tokens * 4
    
    if len(text) <= max_chars:
        return [text]
    
    chunks = []
    start = 0
    
    while start < len(text):
        end = start + max_chars
        
        if end >= len(text):
            chunk = text[start:]
        else:
            # Try to find a good break point (sentence end, then word boundary)
            chunk_candidate = text[start:end]
            
            # Look for sentence end within last 20% of chunk
            search_start = len(chunk_candidate) - int(len(chunk_candidate) * 0.2)
            sentence_end = -1
            
            for i in range(len(chunk_candidate) - 1, search_start, -1):
                if chunk_candidate[i] in '.!?' and i < len(chunk_candidate) - 1:
                    if chunk_candidate[i + 1].isspace():
                        sentence_end = i + 1
                        break
            
            if sentence_end > 0:
                chunk = chunk_candidate[:sentence_end]
            else:
                # Fall back to word boundary
                last_space = chunk_candidate.rfind(' ')
                if last_space > max_chars * 0.7:  # Don't break too early
                    chunk = chunk_candidate[:last_space]
                else:
                    chunk = chunk_candidate
        
        chunks.append(chunk.strip())
        
        # Calculate next start with overlap
        if end >= len(text):
            break
        
        # Find good overlap start
        chunk_len = len(chunk)
        overlap_start = max(0, chunk_len - overlap_chars)
        
        # Adjust to word boundary
        while overlap_start > 0 and not text[start + overlap_start].isspace():
            overlap_start -= 1
        
        start = start + overlap_start if overlap_start > 0 else end
    
    return chunks

# ================================================================== #
# 3. Enhanced File Processing                                       #
# ================================================================== #

def process_uploaded_files(uploaded_files) -> Dict[str, str]:
    """Process multiple uploaded files in parallel with enhanced error handling."""
    def _worker(uploaded_file):
        try:
            file_bytes = uploaded_file.read()
            text = extract_text_from_file(file_bytes, uploaded_file.name)
            
            # Basic text cleaning
            text = _clean_extracted_text(text)
            
            return uploaded_file.name, text
        except Exception as e:
            print(f"Error processing {uploaded_file.name}: {e}")
            return uploaded_file.name, f"Error: Could not extract text from {uploaded_file.name}"

    results = {}
    with cf.ThreadPoolExecutor(max_workers=4) as executor:
        futures = [executor.submit(_worker, f) for f in uploaded_files]
        for future in cf.as_completed(futures):
            filename, text = future.result()
            results[filename] = text
    
    return results

def _clean_extracted_text(text: str) -> str:
    """Clean extracted text for better processing."""
    if not text:
        return ""
    
    # Remove excessive whitespace
    text = re.sub(r'\n\s*\n\s*\n', '\n\n', text)  # Multiple line breaks to double
    text = re.sub(r'[ \t]+', ' ', text)           # Multiple spaces to single
    text = re.sub(r'\r\n', '\n', text)            # Windows line endings
    
    # Remove common OCR artifacts (but keep most punctuation)
    text = re.sub(r'[^\w\s\.,!?@#$%^&*()_+=\[\]{}|;:,.<>?/~`"-]', '', text)
    
    return text.strip()

# ================================================================== #
# 4. Enhanced Vector Database with Hybrid Search                    #
# ================================================================== #

class EnhancedVectorStore:
    """Wrapper for enhanced vector store with ensemble retriever."""
    def __init__(self, vectorstore, ensemble_retriever):
        self.vectorstore = vectorstore
        self.ensemble_retriever = ensemble_retriever
    
    def as_retriever(self, **kwargs):
        """Return the ensemble retriever for hybrid search."""
        return self.ensemble_retriever
    
    def similarity_search(self, query, k=4):
        """Fallback to vector store similarity search."""
        if self.vectorstore:
            return self.vectorstore.similarity_search(query, k=k)
        return []

def create_vector_db(texts: List[str], collection_name: str = "documents_v3"):
    """
    Create enhanced vector database with hybrid search capabilities.
    Combines semantic search (Chroma) with keyword search (BM25).
    """
    if not texts or all(not t.strip() for t in texts):
        print("No valid texts provided for vector DB creation")
        return None

    try:
        print("Creating enhanced vector database with hybrid search...")
        
        # 1. Process and chunk documents using semantic chunking
        all_docs = []
        doc_texts_for_bm25 = []
        
        for doc_idx, text in enumerate(texts):
            if not text.strip():
                continue
            
            # Use hybrid chunking for better quality
            chunks = _hybrid_chunk_text(text, max_tokens=400, overlap_tokens=50)
            
            for chunk_idx, chunk in enumerate(chunks):
                if chunk.strip():
                    # Create Document objects for BM25
                    doc = Document(
                        page_content=chunk,
                        metadata={
                            "source": f"document_{doc_idx}",
                            "chunk": chunk_idx,
                            "doc_id": f"doc_{doc_idx}_chunk_{chunk_idx}",
                            "filename": f"document_{doc_idx}"
                        }
                    )
                    
                    all_docs.append(doc)
                    doc_texts_for_bm25.append(chunk)
        
        if not all_docs:
            print("No valid document chunks created")
            return None
        
        print(f"Created {len(all_docs)} document chunks")
        
        try:
            # 2. Create BM25 retriever for keyword search
            bm25_retriever = BM25Retriever.from_documents(all_docs)
            bm25_retriever.k = 3  # Return top 3 results
        except Exception as e:
            print(f"Failed to create BM25 retriever: {e}")
            bm25_retriever = None
        
        # 3. Create Chroma vector store for semantic search
        try:
            embeddings = OllamaEmbeddings(model="nomic-embed-text")
            
            # Handle existing collection
            persist_directory = ".chromadb_v3"
            client = chromadb.PersistentClient(path=persist_directory)
            
            try:
                existing_collections = client.list_collections()
                if any(c.name == collection_name for c in existing_collections):
                    client.delete_collection(name=collection_name)
                    print(f"Deleted existing collection: {collection_name}")
            except Exception as e:
                print(f"Note: Could not delete collection: {e}")
            
            # Create new vector store
            doc_texts = [doc.page_content for doc in all_docs]
            doc_metadatas = [doc.metadata for doc in all_docs]
            
            vectorstore = Chroma.from_texts(
                texts=doc_texts,
                embedding=embeddings,
                metadatas=doc_metadatas,
                collection_name=collection_name,
                client=client,
                persist_directory=persist_directory
            )
            
        except Exception as e:
            print(f"Failed to create vector store: {e}")
            vectorstore = None
        
        # 4. Create ensemble retriever combining both approaches (if both available)
        vector_retriever = vectorstore.as_retriever(search_kwargs={"k": 3}) if vectorstore else None
        
        if bm25_retriever and vector_retriever:
            try:
                ensemble_retriever = EnsembleRetriever(
                    retrievers=[bm25_retriever, vector_retriever],
                    weights=[0.4, 0.6]  # Favor semantic search slightly
                )
                print("✅ Enhanced vector database with hybrid search created successfully!")
                return EnhancedVectorStore(vectorstore, ensemble_retriever)
            except Exception as e:
                print(f"Failed to create ensemble retriever: {e}")
        
        # Fallback to single retriever
        if vector_retriever:
            print("✅ Vector database created (semantic search only)")
            return vectorstore
        elif bm25_retriever:
            print("✅ BM25 database created (keyword search only)")
            
            # Wrap BM25 retriever to look like a vector store
            class BM25VectorStore:
                def __init__(self, bm25_retriever):
                    self.bm25_retriever = bm25_retriever
                
                def as_retriever(self, **kwargs):
                    return self.bm25_retriever
                
                def similarity_search(self, query, k=4):
                    # BM25 doesn't have similarity_search, so use get_relevant_documents
                    try:
                        return self.bm25_retriever.get_relevant_documents(query)[:k]
                    except:
                        return []
            
            return BM25VectorStore(bm25_retriever)
        else:
            print("❌ Failed to create any retrieval system")
            return None
        
    except Exception as e:
        print(f"Enhanced vector DB creation error: {e}")
        print(f"Error type: {type(e).__name__}")
        return None
